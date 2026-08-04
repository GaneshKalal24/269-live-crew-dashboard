#!/usr/bin/env python3
"""
Parse the 269-ARD Project Schedule Excel export into:
  1. schedule.json      -> per-structure activity lists (for the crew board pins)
  2. schedule_full.json -> the ENTIRE schedule tree (for look-ahead pages)

Usage: python parse_schedule.py <schedule.xlsx>
Re-run this whenever a new schedule is uploaded - one command, no eyeballing.
"""
import sys, json, re
from datetime import datetime, date
import openpyxl

DAY_MONTHS = None

def parse_dt(s):
    """'Mon 4/08/26' or 'Mon 4/08/2026' -> ISO 'YYYY-MM-DD'."""
    if s is None: return None
    s = str(s).strip()
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if not m: return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    try:
        return date(y, mo, d).isoformat()
    except ValueError:
        return None

def indent(name):
    return (len(name) - len(name.lstrip())) // 3

# Maps schedule structure names -> the pin `sched` key used in pins.json
# (schedule names have odd spacing; normalise to the keys the board already uses)
SCHED_KEY_MAP = {
    "TCC Convention Box": "TCC Convention Box",
    "15 - 3CDU-D-15-0001": "15-3CDU-D-15-0001",
    "37-2PTR-B-3701/2 (Small)": "37-2PTR-B-3701/2 (Small)",
    "37-2PTR-B-3703/4/5/6 (Large)": "37-2PTR-B-3703/4/5/6 (Large)",
    "15-3CDU-B-1501 (Large)": "15-3CDU-B-1501 (Large)",
    "12-SGP- B-1103 (Double Skin)": "12-SGP-B-1103 (Double Skin)",
    "11-2CDU-B1101": "11-2CDU-B1101",
    "11-2CDU- B1102": "11-2CDU-B1102",
    "12-SGP-B-1202 ( Round)": "12-SGP-B-1202 (Round)",
    "12-SGP-B-1203 (Round)": "12-SGP-B-1203 (Round)",
    "38-3PTR-B-3804/5/6/7 (Large)": "38-3PTR-B-3804/5/6/7 (Large)",
    "34-BRU-EAST-B-34-0001": "34-BRU-EAST-B-34-0001",
    "38-3PTR-B-3801 (Round)": "38-3PTR-B-3801 (Round)",
    "51-CHD-B-51-0001 (Round)": "51-CHD-B-51-0001 (Round)",
    "38-3PTR- B-3802 (Round)": "38-3PTR-B-3802 (Round)",
    "36-JFT-B-36-0001 (Round)": "36-JFT-B-36-0001 (Round)",
    "Boiler 01": "02-Boilerhouse: Boiler 01",
    "Boiler 02": "02-Boilerhouse: Boiler 02",
    # Top N Tail columns
    "3CDU-D-15-0002/3/4 (Top N Tail)": "3CDU-D-15-0002/3/4",
    "38-3PTR-D-38-3813 (Top N Tail)": "38-3PTR-D-38-3813",
    "11-2CDU-D-11-0002/3/4 (Top N Tail)": "11-2CDU-D-11-0002/3/4",
    "27-USGP-D-27-0002/3/4 (Top N Tail)": "27-USGP-D-27-0002/3/4",
    "38-3PTR-D-38-3302 (Top N Tail)": "38-3PTR-D-38-3302",
    "36-JFT-D-36-3301 (Top N Tail)": "36-JFT-D-36-3301",
    "45-ALKY-DA-45-0003 (Top N Tail)": "45-ALKY-DA-45-0003",
    "45-ALKY-DA-45-0004 (Top N Tail)": "45-ALKY-DA-45-0004",
    "51-CHD-D-51-0003 (Top N Tail)": "51-CHD-D-51-0003",
    "51-CHD-D-51-0005 (Top N Tail)": "51-CHD-D-51-0005",
}

ACTIVITY_NAMES = {
    "Erect Scaffold","Shrink Wrap","Flooring Encapsulation","Smoke Test",
    "Remove Asbestos","Clearance + Remove Encapsulation","Clearance",
    "Dismantle Scaffold","Remove Construction Joints - Asbestos"
}

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # ---- 1. Full tree (for look-ahead pages) ----
    full = []
    for row in rows[1:]:
        if row[0] is None: continue
        name = str(row[0])
        st = name.strip()
        start = parse_dt(row[2]); finish = parse_dt(row[3])
        pct = row[4]
        try: pct = round(float(pct)*100) if pct is not None else 0
        except (TypeError, ValueError): pct = 0
        dur = str(row[1]).strip() if row[1] else ""
        full.append({"name": st, "level": indent(name), "start": start,
                     "finish": finish, "pct": pct, "duration": dur})

    # ---- 2. Per-structure activities for the board ----
    schedule = {}
    current_key = None
    for row in rows[1:]:
        if row[0] is None: continue
        name = str(row[0]); st = name.strip()
        # normalise activity name that may have "(Potential Double Skin)..." suffix
        act_base = st
        if st.startswith("Remove Asbestos"): act_base = "Remove Asbestos"
        # is this a structure header we track?
        if st in SCHED_KEY_MAP:
            current_key = SCHED_KEY_MAP[st]
            schedule[current_key] = []
            continue
        # is this an activity row under the current structure?
        if current_key and act_base in ACTIVITY_NAMES:
            s = parse_dt(row[2]); f = parse_dt(row[3])
            pct = row[4]
            try: pct = round(float(pct)*100) if pct is not None else 0
            except (TypeError, ValueError): pct = 0
            if s and f:
                schedule[current_key].append({"a": act_base, "s": s, "f": f, "p": pct})
        # a new non-activity, non-tracked header ends the current structure
        elif current_key and indent(name) <= 5 and st not in SCHED_KEY_MAP:
            current_key = None

    # structures with no dates yet (kept so pins still exist)
    for extra in ["38-0006 Naphtha Splitter","38-0004 Depentaniser"]:
        schedule.setdefault(extra, [])

    with open('schedule.json','w') as fp:
        json.dump(schedule, fp, separators=(',',':'))
    with open('schedule_full.json','w') as fp:
        json.dump(full, fp, separators=(',',':'))

    print(f"schedule.json: {len(schedule)} structures")
    print(f"schedule_full.json: {len(full)} rows")
    # quick sanity
    live = [(k,a['a'],a['p']) for k,acts in schedule.items() for a in acts if 0 < a['p'] < 100]
    print(f"in-progress activities: {len(live)}")
    for k,a,p in live[:12]:
        print(f"  {k} · {a} · {p}%")

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else 'Project_Schedule_4_08_2026.xlsx')
